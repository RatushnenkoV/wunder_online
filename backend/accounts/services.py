import csv
import io
import secrets
import string

from openpyxl import load_workbook

from .models import User


def generate_password(length=8):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def create_user_with_temp_password(first_name, last_name, roles=None, email='', phone=''):
    password = generate_password()
    user = User(
        first_name=first_name,
        last_name=last_name,
        email=email,
        phone=phone,
        must_change_password=True,
        temp_password=password,
    )
    if roles:
        for role in roles:
            if role == 'admin':
                user.is_admin = True
                user.is_staff = True
            elif role == 'teacher':
                user.is_teacher = True
            elif role == 'parent':
                user.is_parent = True
            elif role == 'student':
                user.is_student = True
    user.set_password(password)
    user.save()
    return user


def reset_user_password(user):
    password = generate_password()
    user.set_password(password)
    user.temp_password = password
    user.must_change_password = True
    user.save()
    return password


def parse_import_file(file):
    """Parse uploaded CSV or XLSX file into list of dicts."""
    name = file.name.lower()
    if name.endswith('.csv'):
        return _parse_csv(file)
    elif name.endswith('.xlsx'):
        return _parse_xlsx(file)
    else:
        raise ValueError('Поддерживаются только CSV и XLSX файлы')


def _parse_csv(file):
    content = file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content), delimiter=';')
    return list(reader)


def _parse_xlsx(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        result.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
    return result


def import_users(file):
    """
    Import users from file. Expected columns:
    фамилия, имя, роли (comma-separated: admin,teacher,parent,student), email, телефон
    """
    rows = parse_import_file(file)
    created = []
    errors = []
    for i, row in enumerate(rows, start=2):
        try:
            last_name = row.get('фамилия', '').strip()
            first_name = row.get('имя', '').strip()
            roles_str = row.get('роли', row.get('роль', '')).strip()
            email = row.get('email', row.get('почта', '')).strip()
            phone = row.get('телефон', '').strip()

            if not last_name or not first_name:
                errors.append(f'Строка {i}: имя и фамилия обязательны')
                continue

            roles = [r.strip() for r in roles_str.split(',') if r.strip()] if roles_str else []
            user = create_user_with_temp_password(first_name, last_name, roles, email, phone)
            created.append(user)
        except Exception as e:
            errors.append(f'Строка {i}: {str(e)}')

    return created, errors
