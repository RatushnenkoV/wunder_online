from datetime import timedelta, date as date_type
from collections import Counter

from accounts.services import parse_import_file
from school.models import ScheduleLesson
from .models import Holiday, SchoolBreak, Topic


def _get_skip_dates():
    """
    Return a set of dates to skip during scheduling:
    individual holidays + all days within school breaks.
    """
    skip = set(Holiday.objects.values_list('date', flat=True))
    for sb in SchoolBreak.objects.all():
        d = sb.start_date
        while d <= sb.end_date:
            skip.add(d)
            d += timedelta(days=1)
    return skip


def get_schedule_info(ctp):
    """
    Get schedule info for a CTP: which weekdays and how many lessons per day.
    Returns dict: {weekday_number(0=Mon): lessons_count, ...}

    ScheduleLesson.weekday uses 1=Mon..5=Fri, we convert to Python's 0=Mon..4=Fri.
    """
    lessons = ScheduleLesson.objects.filter(
        school_class=ctp.school_class,
        subject=ctp.subject,
    )

    weekday_counts = Counter()
    for lesson in lessons:
        # Convert 1-based weekday to 0-based (Python convention)
        weekday_counts[lesson.weekday - 1] += 1

    return dict(weekday_counts)


def get_required_lessons_count(ctp, start_date, end_date=None):
    """
    Calculate how many lessons are required based on schedule,
    from start_date to end_date (defaults to May 31 of school year).
    """
    if end_date is None:
        # End of school year: May 31
        if start_date.month >= 9:
            end_date = start_date.replace(year=start_date.year + 1, month=5, day=31)
        else:
            end_date = start_date.replace(month=5, day=31)

    schedule_info = get_schedule_info(ctp)
    if not schedule_info:
        return 0

    skip_dates = _get_skip_dates()

    count = 0
    current_date = start_date
    while current_date <= end_date:
        weekday = current_date.weekday()
        if weekday in schedule_info and current_date not in skip_dates:
            count += schedule_info[weekday]
        current_date += timedelta(days=1)

    return count


def autofill_dates(ctp, start_date, weekdays=None, lessons_per_day=1,
                   start_from_topic_id=None, use_schedule=True):
    """
    Auto-fill dates for topics in a CTP.

    Args:
        ctp: CTP instance
        start_date: date to start from
        weekdays: list of ints (0=Mon, 1=Tue, ..., 6=Sun) - used if use_schedule=False
        lessons_per_day: how many lessons per day - used if use_schedule=False
        start_from_topic_id: if set, only fill dates starting from this topic
        use_schedule: if True, use schedule data for weekdays and lessons_per_day
    """
    skip_dates = _get_skip_dates()

    # Get schedule-based weekday info: {weekday: lessons_count}
    if use_schedule:
        schedule_info = get_schedule_info(ctp)
        if not schedule_info:
            return
    else:
        # Legacy mode: use provided weekdays with uniform lessons_per_day
        if not weekdays:
            return
        schedule_info = {d: lessons_per_day for d in weekdays}

    topics = list(ctp.topics.order_by('order'))

    if start_from_topic_id:
        start_idx = None
        for i, t in enumerate(topics):
            if t.id == start_from_topic_id:
                start_idx = i
                break
        if start_idx is None:
            return
        topics = topics[start_idx:]

    current_date = start_date
    topic_idx = 0

    while topic_idx < len(topics):
        weekday = current_date.weekday()
        if weekday in schedule_info and current_date not in skip_dates:
            for _ in range(schedule_info[weekday]):
                if topic_idx >= len(topics):
                    break
                topics[topic_idx].date = current_date
                topics[topic_idx].save(update_fields=['date'])
                topic_idx += 1
        current_date += timedelta(days=1)


def _parse_ktp_xlsx(file):
    """
    Parse KTP XLSX with header on row 3 (1-indexed).
    Finds the header row by looking for 'тема урока' or 'тема', then reads data below.
    Falls back to standard parse if header not found on row 3.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row: look for row containing 'тема урока' or 'тема'
    header_idx = None
    for i, row in enumerate(rows):
        row_lower = [str(v).strip().lower() if v else '' for v in row]
        if 'тема урока' in row_lower or 'тема' in row_lower:
            header_idx = i
            break

    if header_idx is None:
        # Fallback: use first row as header
        header_idx = 0

    headers = [str(h).strip().lower() if h else '' for h in rows[header_idx]]
    result = []
    for row in rows[header_idx + 1:]:
        if all(cell is None for cell in row):
            continue
        result.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
    return result


def import_topics(ctp, file):
    """
    Import topics from CSV/XLSX.
    Supports standard format (header on row 1) and KTP format (header on row 3).
    Columns: тема урока, дата урока, домашнее задание, комментарии,
             cсылки на самообуч, дополнительные ресурсы,
             индивид. папка ученика, ксп, ссылка на презентацию
    """
    name = getattr(file, 'name', '').lower()
    if name.endswith('.xlsx'):
        rows = _parse_ktp_xlsx(file)
    else:
        rows = parse_import_file(file)

    last_order = ctp.topics.count()
    created = []
    errors = []

    for i, row in enumerate(rows, start=2):
        # Title: try several column name variants
        title = (row.get('тема урока') or row.get('тема') or row.get('название') or '').strip()
        if not title:
            errors.append(f'Строка {i}: название темы обязательно')
            continue

        # Date
        raw_date = (row.get('дата урока') or row.get('дата') or '').strip()
        parsed_date = None
        if raw_date:
            for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                try:
                    from datetime import datetime
                    parsed_date = datetime.strptime(raw_date.split(' ')[0], fmt).date()
                    break
                except ValueError:
                    continue

        topic = Topic.objects.create(
            ctp=ctp,
            order=last_order,
            title=title,
            date=parsed_date,
            homework=(row.get('домашнее задание') or row.get('домашнее задание ') or '').strip(),
            comments=(row.get('комментарии') or '').strip(),
            self_study_links=(row.get('cсылки на самообуч') or row.get('ссылки на самообуч') or '').strip(),
            additional_resources=(row.get('дополнительные ресурсы') or '').strip(),
            individual_folder=(row.get('индивид. папка ученика') or row.get('индивидуальная папка ученика') or '').strip(),
            ksp=(row.get('ксп') or row.get('кsp') or '').strip(),
            presentation_link=(row.get('ссылка на презентацию') or '').strip(),
        )
        created.append(topic)
        last_order += 1

    return created, errors


# ── Bulk import helpers ────────────────────────────────────────────────────────

def _match_subject(name, all_subjects):
    """Try to match a sheet name to a Subject. Returns Subject or None."""
    import re
    clean = re.sub(r'\s*\d+\s*(?:ч\.|ч|часов?)\.?\s*$', '', name).strip()
    # Exact
    for s in all_subjects:
        if s.name.lower() == clean.lower():
            return s
    # One contains the other
    for s in all_subjects:
        sn = s.name.lower()
        cn = clean.lower()
        if cn.startswith(sn) or sn.startswith(cn) or sn in cn or cn in sn:
            return s
    return None


def preview_ktp_import(file):
    """
    Analyze xlsx file for bulk KTP import.
    Returns detected class name + per-sheet analysis.
    """
    import re
    from openpyxl import load_workbook
    from school.models import Subject

    wb = load_workbook(file, read_only=True)

    # Detect class from filename
    filename = getattr(file, 'name', '')
    detected_class_name = None
    m = re.search(r'(\d+)\s*(?:класс|кл\.?)', filename, re.IGNORECASE)
    if m:
        detected_class_name = f'{m.group(1)} класс'

    all_subjects = list(Subject.objects.all())

    sheets = []
    for ws in wb.worksheets:
        # Count topics: find header row, count non-empty rows after it
        sample = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        header_idx = None
        for i, row in enumerate(sample):
            row_lower = [str(v).strip().lower() if v else '' for v in row]
            if 'тема урока' in row_lower or 'тема' in row_lower:
                header_idx = i
                break

        if header_idx is not None:
            all_rows = list(ws.iter_rows(values_only=True))
            topics_count = sum(
                1 for row in all_rows[header_idx + 1:]
                if any(v for v in row[:8])
            )
        else:
            topics_count = 0

        matched = _match_subject(ws.title, all_subjects)
        sheets.append({
            'sheet_name': ws.title,
            'matched_subject_id': matched.id if matched else None,
            'matched_subject_name': matched.name if matched else None,
            'topics_count': topics_count,
        })

    wb.close()
    return {
        'detected_class_name': detected_class_name,
        'sheets': sheets,
    }


def _parse_row_to_topic_fields(row_dict):
    """Extract topic field values from a row dict (lowercased keys)."""
    from datetime import datetime

    title = (row_dict.get('тема урока') or row_dict.get('тема') or row_dict.get('название') or '').strip()

    raw_date = (row_dict.get('дата урока') or row_dict.get('дата') or '').strip()
    parsed_date = None
    if raw_date:
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                parsed_date = datetime.strptime(raw_date.split(' ')[0], fmt).date()
                break
            except ValueError:
                continue

    return {
        'title': title,
        'date': parsed_date,
        'homework': (row_dict.get('домашнее задание') or row_dict.get('домашнее задание ') or '').strip(),
        'comments': (row_dict.get('комментарии') or '').strip(),
        'self_study_links': (row_dict.get('cсылки на самообуч') or row_dict.get('ссылки на самообуч') or '').strip(),
        'additional_resources': (row_dict.get('дополнительные ресурсы') or '').strip(),
        'individual_folder': (row_dict.get('индивид. папка ученика') or row_dict.get('индивидуальная папка ученика') or '').strip(),
        'ksp': (row_dict.get('ксп') or row_dict.get('кsp') or '').strip(),
        'presentation_link': (row_dict.get('ссылка на презентацию') or '').strip(),
    }


def import_ktp_from_sheet(file, sheet_name, class_id, subject_id, teacher):
    """
    Import topics from a specific xlsx sheet into a new CTP.
    Returns (ctp, created_count, errors).
    """
    from openpyxl import load_workbook
    from .models import CTP, Topic

    wb = load_workbook(file, read_only=True)

    ws = next((s for s in wb.worksheets if s.title == sheet_name), None)
    if ws is None:
        raise ValueError(f'Лист "{sheet_name}" не найден в файле')

    rows_all = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row
    header_idx = None
    for i, row in enumerate(rows_all):
        row_lower = [str(v).strip().lower() if v else '' for v in row]
        if 'тема урока' in row_lower or 'тема' in row_lower:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f'В листе "{sheet_name}" не найдена строка заголовков')

    headers = [str(h).strip().lower() if h else '' for h in rows_all[header_idx]]

    ctp = CTP.objects.create(
        teacher=teacher,
        school_class_id=class_id,
        subject_id=subject_id,
        is_public=True,
    )

    created_count = 0
    errors = []
    order = 0

    for row in rows_all[header_idx + 1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = dict(zip(headers, [str(v).strip() if v is not None else '' for v in row]))
        fields = _parse_row_to_topic_fields(row_dict)
        if not fields['title']:
            continue
        Topic.objects.create(ctp=ctp, order=order, **fields)
        created_count += 1
        order += 1

    return ctp, created_count, errors
