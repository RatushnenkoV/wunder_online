import datetime
import logging

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status

from accounts.permissions import PasswordChanged
from accounts.permissions import IsAdmin
from .models import SchoolEvent
from .serializers import SchoolEventSerializer

logger = logging.getLogger(__name__)


def _is_staff(user):
    return user and (user.is_admin or user.is_teacher or user.is_spps)


def _can_edit(user):
    return user and (user.is_admin or user.is_teacher)


@api_view(['GET', 'POST'])
@permission_classes([PasswordChanged])
def event_list_create(request):
    if not _is_staff(request.user):
        return Response({'detail': 'Нет доступа.'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        qs = SchoolEvent.objects.all()
        date_after = request.query_params.get('date_after')
        date_before = request.query_params.get('date_before')
        event_type = request.query_params.get('event_type')
        if date_after:
            qs = qs.filter(date_start__gte=date_after)
        if date_before:
            qs = qs.filter(date_start__lte=date_before)
        if event_type:
            qs = qs.filter(event_type=event_type)
        return Response(SchoolEventSerializer(qs, many=True).data)

    if not _can_edit(request.user):
        return Response({'detail': 'Нет доступа.'}, status=status.HTTP_403_FORBIDDEN)

    serializer = SchoolEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(created_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([PasswordChanged])
def event_detail(request, pk):
    if not _is_staff(request.user):
        return Response({'detail': 'Нет доступа.'}, status=status.HTTP_403_FORBIDDEN)

    try:
        event = SchoolEvent.objects.get(pk=pk)
    except SchoolEvent.DoesNotExist:
        return Response({'detail': 'Не найдено.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(SchoolEventSerializer(event).data)

    if not _can_edit(request.user):
        return Response({'detail': 'Нет доступа.'}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'DELETE':
        if not request.user.is_admin:
            return Response({'detail': 'Нет доступа.'}, status=status.HTTP_403_FORBIDDEN)
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = SchoolEventSerializer(event, data=request.data, partial=(request.method == 'PATCH'))
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


EVENT_TYPE_MAP = {
    'праздник': 'holiday',
    'тимбилдинг': 'teambuilding',
    'метапредметный проект': 'meta_subject',
    'межпредметный проект': 'cross_subject',
    'предметный проект': 'subject',
    'обучение': 'training',
    'профориентация': 'career_guidance',
}


def _map_event_type(value):
    if not value:
        return 'other'
    # ' '.join(split()) normalises all whitespace variants incl. non-breaking spaces (\xa0)
    normalized = ' '.join(str(value).strip().lower().split())
    return EVENT_TYPE_MAP.get(normalized, 'other')


def _map_approved(value):
    if value is True:
        return 'yes'
    if value is False:
        return 'no'
    if value and 'перенес' in str(value).lower():
        return 'rescheduled'
    return 'pending'


def _parse_date_cell(cell_value):
    """Returns (date_start, date_end). date_end may be None."""
    if isinstance(cell_value, datetime.datetime):
        return cell_value.date(), None
    if isinstance(cell_value, datetime.date):
        return cell_value, None
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
        if ' - ' in cell_value:
            parts = cell_value.split(' - ')
            try:
                d1 = datetime.datetime.strptime(parts[0].strip(), '%d.%m.%Y').date()
                d2 = datetime.datetime.strptime(parts[1].strip(), '%d.%m.%Y').date()
                return d1, d2
            except ValueError:
                pass
        try:
            return datetime.datetime.strptime(cell_value, '%d.%m.%Y').date(), None
        except ValueError:
            pass
    return None, None


def _join(*args):
    return ', '.join(str(a).strip() for a in args if a and str(a).strip())


@api_view(['GET', 'POST'])
@permission_classes([IsAdmin, PasswordChanged])
@parser_classes([MultiPartParser, FormParser])
def event_import(request):
    """
    GET: Returns the current count of events (used before showing import dialog).
    POST: Imports events from xlsx file.
          params: file (xlsx), replace ('true'/'false')
    """
    if request.method == 'GET':
        return Response({'count': SchoolEvent.objects.count()})

    file = request.FILES.get('file')
    if not file:
        return Response({'detail': 'Файл не передан.'}, status=status.HTTP_400_BAD_REQUEST)

    replace = request.data.get('replace', 'false').lower() == 'true'

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return Response({'detail': f'Ошибка чтения файла: {e}'}, status=status.HTTP_400_BAD_REQUEST)

    if 'План' not in wb.sheetnames:
        available = ', '.join(wb.sheetnames)
        logger.error('events import: sheet "План" not found. Available: %s', available)
        return Response(
            {'detail': f'Лист "План" не найден. Доступные листы: {available}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    ws = wb['План']
    events_to_create = []
    skipped = 0
    skipped_reasons: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        row_num = i + 3
        # columns (0-indexed): 0=month, 1=date, 2=time, 3=class1, 4=class2,
        # 5=organizer1, 6=organizer2, 7=description, 8=responsible, 9=helper,
        # 10=event_type, 11=approved, 12=cost, 13=status
        date_val = row[1] if len(row) > 1 else None
        description = row[7] if len(row) > 7 else None

        if not description and not date_val:
            skipped += 1
            continue

        if not description:
            skipped += 1
            skipped_reasons.append(f'строка {row_num}: нет описания (дата={date_val})')
            continue

        description = str(description).strip()
        if not description:
            skipped += 1
            continue

        if not date_val:
            skipped += 1
            skipped_reasons.append(f'строка {row_num}: нет даты (описание={description[:40]})')
            continue

        date_start, date_end = _parse_date_cell(date_val)
        if not date_start:
            skipped += 1
            skipped_reasons.append(f'строка {row_num}: не удалось распарсить дату {date_val!r} (описание={description[:40]})')
            continue

        time_note = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        target_classes = _join(
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
        )
        organizers = _join(
            row[5] if len(row) > 5 else None,
            row[6] if len(row) > 6 else None,
        )
        responsible = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        helper = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        event_type = _map_event_type(row[10] if len(row) > 10 else None)
        approved = _map_approved(row[11] if len(row) > 11 else None)
        cost_raw = row[12] if len(row) > 12 else None
        cost = str(cost_raw).strip() if cost_raw is not None else ''
        status_val = str(row[13]).strip() if len(row) > 13 and row[13] else ''

        events_to_create.append(SchoolEvent(
            date_start=date_start,
            date_end=date_end,
            time_note=time_note,
            target_classes=target_classes,
            organizers=organizers,
            description=description,
            responsible=responsible,
            helper=helper,
            event_type=event_type,
            approved=approved,
            cost=cost,
            status=status_val,
            created_by=request.user,
        ))

    logger.info(
        'events import by %s: parsed %d events, skipped %d rows, replace=%s',
        request.user, len(events_to_create), skipped, replace,
    )
    if skipped_reasons:
        logger.warning('events import skipped rows:\n%s', '\n'.join(skipped_reasons))

    try:
        if replace:
            deleted_count, _ = SchoolEvent.objects.all().delete()
            logger.info('events import: deleted %d existing events', deleted_count)

        SchoolEvent.objects.bulk_create(events_to_create)
    except Exception as exc:
        logger.exception('events import: DB error during bulk_create')
        return Response({'detail': f'Ошибка записи в базу данных: {exc}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({
        'created': len(events_to_create),
        'skipped': skipped,
        'replaced': replace,
        'skipped_details': skipped_reasons[:20],  # first 20 reasons for debugging
    }, status=status.HTTP_201_CREATED)
