import datetime

from django.contrib.auth.hashers import make_password
from accounts.services import create_user_with_temp_password, generate_password, parse_import_file
from accounts.models import User
from .models import GradeLevel, SchoolClass, StudentProfile, ParentProfile


def import_classes(file):
    """
    Import classes with students and parents from CSV/XLSX.
    Expected columns:
    параллель, буква, фамилия_ученика, имя_ученика, email_ученика, телефон_ученика,
    фамилия_родителя1, имя_родителя1, email_родителя1, телефон_родителя1,
    фамилия_родителя2, имя_родителя2, email_родителя2, телефон_родителя2
    """
    rows = parse_import_file(file)
    created_students = []
    created_parents = []
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            grade_num = int(row.get('параллель', 0))
            letter = _normalize_class_letter(row.get('буква', ''))
            student_last = row.get('фамилия_ученика', '').strip()
            student_first = row.get('имя_ученика', '').strip()

            if not grade_num or not letter or not student_last or not student_first:
                errors.append(f'Строка {i}: параллель, буква, фамилия и имя ученика обязательны')
                continue

            grade, _ = GradeLevel.objects.get_or_create(number=grade_num)
            school_class, _ = SchoolClass.objects.get_or_create(grade_level=grade, letter=letter)

            student_user = create_user_with_temp_password(
                student_first, student_last, ['student'],
                row.get('email_ученика', '').strip(),
                row.get('телефон_ученика', '').strip(),
            )
            student_profile = StudentProfile.objects.create(user=student_user, school_class=school_class)
            created_students.append(student_user)

            for suffix in ['1', '2']:
                p_last = row.get(f'фамилия_родителя{suffix}', '').strip()
                p_first = row.get(f'имя_родителя{suffix}', '').strip()
                if p_last and p_first:
                    parent_user = create_user_with_temp_password(
                        p_first, p_last, ['parent'],
                        row.get(f'email_родителя{suffix}', '').strip(),
                        row.get(f'телефон_родителя{suffix}', '').strip(),
                    )
                    parent_profile, _ = ParentProfile.objects.get_or_create(user=parent_user)
                    parent_profile.children.add(student_profile)
                    created_parents.append(parent_user)

        except Exception as e:
            errors.append(f'Строка {i}: {str(e)}')

    return {
        'students': created_students,
        'parents': created_parents,
        'errors': errors,
    }


def _normalize_class_letter(letter: str) -> str:
    """Normalize class letter: uppercase, then map Latin lookalikes to Cyrillic."""
    letter = letter.strip().upper()
    # Latin uppercase lookalikes → Cyrillic uppercase
    latin_to_cyrillic = {
        'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н',
        'K': 'К', 'M': 'М', 'O': 'О', 'P': 'Р', 'T': 'Т',
        'X': 'Х', 'Y': 'У',
    }
    return latin_to_cyrillic.get(letter, letter)


def _parse_class_str(class_str):
    """Parse '1 а', '2 б', '9 А' → (grade_num, letter). Normalises to uppercase Cyrillic."""
    parts = class_str.strip().split()
    if len(parts) < 2:
        raise ValueError(f'Не удалось распознать класс: "{class_str}"')
    grade_num = int(parts[0])
    letter = _normalize_class_letter(parts[1])
    return grade_num, letter


_IMPORT_CHUNK = 50


def _gen_username(first: str, last: str, existing: set, reserved: set) -> str:
    """Generate a unique username in-memory, without extra DB queries."""
    base = f'{first}_{last}'.lower()
    username = base
    counter = 1
    while username in existing or username in reserved:
        username = f'{base}_{counter}'
        counter += 1
    reserved.add(username)
    return username


def import_students_from_excel_streaming(file):
    """
    Streaming import using bulk DB operations.
    Yields dicts:
      {'type': 'start',    'total': N}
      {'type': 'progress', 'processed': N, 'total': N, 'created': N, 'updated': N}
      {'type': 'done',     'created': N, 'updated': N, 'errors': [...]}

    Strategy:
      1. Parse all rows from Excel (no DB).
      2. Bulk get/create GradeLevels and SchoolClasses (~4 queries total).
      3. Bulk fetch all existing students by file_number and by name (2 queries).
      4. Pre-load all usernames for in-memory unique generation (1 query).
      5. Process rows in chunks of _IMPORT_CHUNK:
           bulk_update for existing + bulk_create for new (4 queries/chunk).
      Total queries: O(chunks) instead of O(rows).
    """
    from openpyxl import load_workbook

    wb = load_workbook(file)
    errors = []

    # --- Phase 1: Parse all rows (no DB) ---
    parsed = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) <= 1:
            continue
        for row_idx, row in enumerate(all_rows[1:], start=2):
            fio_raw = row[2]
            if not fio_raw:
                continue
            fio = str(fio_raw).strip()
            if not fio:
                continue

            class_raw = row[1]
            if not class_raw:
                errors.append(f'Лист «{sheet_name}», строка {row_idx}: пустое поле «Класс»')
                continue

            birth_date_raw = row[3]
            personal_file_number = str(row[4]).strip() if row[4] is not None else ''

            try:
                grade_num, letter = _parse_class_str(str(class_raw))
            except ValueError as exc:
                errors.append(f'Лист «{sheet_name}», строка {row_idx}: {exc}')
                continue

            parts = fio.split()
            if len(parts) < 2:
                errors.append(
                    f'Лист «{sheet_name}», строка {row_idx}: '
                    f'ФИО должно содержать как минимум фамилию и имя — «{fio}»'
                )
                continue

            birth_date = None
            if isinstance(birth_date_raw, datetime.datetime):
                birth_date = birth_date_raw.date()
            elif isinstance(birth_date_raw, datetime.date):
                birth_date = birth_date_raw

            parsed.append({
                'last_name': parts[0],
                'first_name': ' '.join(parts[1:]),
                'birth_date': birth_date,
                'grade_num': grade_num,
                'letter': letter,
                'personal_file_number': personal_file_number,
                'sheet': sheet_name,
                'row_idx': row_idx,
            })

    total = len(parsed)
    yield {'type': 'start', 'total': total}

    if not parsed:
        yield {'type': 'done', 'created': 0, 'updated': 0, 'errors': errors}
        return

    # --- Phase 2: Bulk resolve GradeLevels (≤2 queries) ---
    grade_nums = {r['grade_num'] for r in parsed}
    existing_grades = {g.number: g for g in GradeLevel.objects.filter(number__in=grade_nums)}
    new_grade_nums = grade_nums - set(existing_grades)
    if new_grade_nums:
        GradeLevel.objects.bulk_create(
            [GradeLevel(number=n) for n in new_grade_nums],
            ignore_conflicts=True,
        )
        for g in GradeLevel.objects.filter(number__in=new_grade_nums):
            existing_grades[g.number] = g

    # --- Phase 3: Bulk resolve SchoolClasses (≤2 queries) ---
    needed_pairs = {(r['grade_num'], r['letter']) for r in parsed}
    classes_qs = SchoolClass.objects.filter(
        grade_level__in=existing_grades.values()
    ).select_related('grade_level')
    class_cache = {(sc.grade_level.number, sc.letter): sc for sc in classes_qs}

    missing_pairs = needed_pairs - set(class_cache)
    if missing_pairs:
        SchoolClass.objects.bulk_create(
            [SchoolClass(grade_level=existing_grades[gn], letter=lt) for gn, lt in missing_pairs]
        )
        # Reload cache for grades that got new classes
        new_grade_ids = {existing_grades[gn].pk for gn, _ in missing_pairs}
        for sc in SchoolClass.objects.filter(
            grade_level_id__in=new_grade_ids
        ).select_related('grade_level'):
            class_cache[(sc.grade_level.number, sc.letter)] = sc

    # --- Phase 4: Bulk lookup existing students (2 queries) ---
    file_numbers = {r['personal_file_number'] for r in parsed if r['personal_file_number']}
    profiles_by_file_num: dict = {}
    if file_numbers:
        for p in StudentProfile.objects.filter(
            personal_file_number__in=file_numbers
        ).select_related('user'):
            profiles_by_file_num[p.personal_file_number] = p

    last_names = {r['last_name'] for r in parsed}
    users_by_name: dict = {}
    for u in User.objects.filter(
        last_name__in=last_names, is_student=True
    ).select_related('student_profile'):
        if hasattr(u, 'student_profile'):
            users_by_name[(u.last_name, u.first_name)] = u

    # Pre-load all existing usernames for unique generation (1 query)
    existing_usernames: set = set(User.objects.values_list('username', flat=True))
    reserved_usernames: set = set()

    # --- Phase 5: Process in chunks ---
    created_count = 0
    updated_count = 0

    for chunk_start in range(0, len(parsed), _IMPORT_CHUNK):
        chunk = parsed[chunk_start:chunk_start + _IMPORT_CHUNK]

        users_to_update: list = []
        profiles_to_update: list = []
        users_to_create: list = []
        create_meta: list = []   # (row_dict, school_class)
        chunk_updated = 0

        for r in chunk:
            school_class = class_cache.get((r['grade_num'], r['letter']))
            if not school_class:
                errors.append(
                    f"Лист «{r['sheet']}», строка {r['row_idx']}: "
                    f"класс {r['grade_num']} {r['letter']} не найден после создания"
                )
                continue

            fn = r['personal_file_number']
            if fn and fn in profiles_by_file_num:
                # Update by personal file number
                profile = profiles_by_file_num[fn]
                user = profile.user
                user.last_name = r['last_name']
                user.first_name = r['first_name']
                user.birth_date = r['birth_date']
                users_to_update.append(user)
                profile.school_class = school_class
                profiles_to_update.append(profile)
                chunk_updated += 1
            elif (r['last_name'], r['first_name']) in users_by_name:
                # Update by name
                user = users_by_name[(r['last_name'], r['first_name'])]
                user.birth_date = r['birth_date']
                users_to_update.append(user)
                profile = getattr(user, 'student_profile', None)
                if profile:
                    profile.school_class = school_class
                    if fn:
                        profile.personal_file_number = fn
                    profiles_to_update.append(profile)
                chunk_updated += 1
            else:
                # Create new student
                password = generate_password()
                username = _gen_username(
                    r['first_name'], r['last_name'],
                    existing_usernames, reserved_usernames,
                )
                user = User(
                    username=username,
                    first_name=r['first_name'],
                    last_name=r['last_name'],
                    birth_date=r['birth_date'],
                    is_student=True,
                    must_change_password=True,
                    temp_password=password,
                    # MD5 для скорости: пользователь обязан сменить пароль при первом входе
                    password=make_password(password, hasher='md5'),
                )
                users_to_create.append(user)
                create_meta.append((r, school_class))

        # Bulk update existing
        if users_to_update:
            User.objects.bulk_update(users_to_update, ['first_name', 'last_name', 'birth_date'])
        if profiles_to_update:
            StudentProfile.objects.bulk_update(
                profiles_to_update, ['school_class', 'personal_file_number']
            )

        # Bulk create new
        if users_to_create:
            created_users = User.objects.bulk_create(users_to_create)
            StudentProfile.objects.bulk_create([
                StudentProfile(
                    user=u,
                    school_class=meta[1],
                    personal_file_number=meta[0]['personal_file_number'],
                )
                for u, meta in zip(created_users, create_meta)
            ])

        created_count += len(users_to_create)
        updated_count += chunk_updated
        processed = min(chunk_start + _IMPORT_CHUNK, total)
        yield {
            'type': 'progress',
            'processed': processed,
            'total': total,
            'created': created_count,
            'updated': updated_count,
        }

    yield {'type': 'done', 'created': created_count, 'updated': updated_count, 'errors': errors}


def import_students_from_excel(file):
    """
    Import students from the school's Excel file ("Ученики по классам").
    Each sheet represents one class. Expected columns (positional):
      0: № | 1: Класс | 2: ФИО | 3: Дата рождения | 4: Номер Л/Д

    Logic:
      - If a student with the same personal_file_number exists → update.
      - Else if a student with the same last_name+first_name exists in the same class → update.
      - Otherwise → create new student.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file)
    created_count = 0
    updated_count = 0
    errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) <= 1:
            continue

        for row_idx, row in enumerate(all_rows[1:], start=2):
            fio_raw = row[2]
            if not fio_raw:
                continue
            fio = str(fio_raw).strip()
            if not fio:
                continue

            class_raw = row[1]
            if not class_raw:
                errors.append(f'Лист «{sheet_name}», строка {row_idx}: пустое поле «Класс»')
                continue

            birth_date_raw = row[3]
            personal_file_number = str(row[4]).strip() if row[4] is not None else ''

            # Parse class
            try:
                grade_num, letter = _parse_class_str(str(class_raw))
            except ValueError as exc:
                errors.append(f'Лист «{sheet_name}», строка {row_idx}: {exc}')
                continue

            # Parse FIO: first word = last_name, rest = first_name
            parts = fio.split()
            if len(parts) < 2:
                errors.append(f'Лист «{sheet_name}», строка {row_idx}: ФИО должно содержать как минимум фамилию и имя — «{fio}»')
                continue
            last_name = parts[0]
            first_name = ' '.join(parts[1:])

            # Parse birth_date
            birth_date = None
            if isinstance(birth_date_raw, datetime.datetime):
                birth_date = birth_date_raw.date()
            elif isinstance(birth_date_raw, datetime.date):
                birth_date = birth_date_raw

            try:
                grade, _ = GradeLevel.objects.get_or_create(number=grade_num)
                school_class, _ = SchoolClass.objects.get_or_create(grade_level=grade, letter=letter)

                # Try to find by personal_file_number first
                existing_profile = None
                if personal_file_number:
                    existing_profile = StudentProfile.objects.filter(
                        personal_file_number=personal_file_number
                    ).select_related('user').first()

                if existing_profile:
                    user = existing_profile.user
                    user.last_name = last_name
                    user.first_name = first_name
                    user.birth_date = birth_date
                    user.save()
                    existing_profile.school_class = school_class
                    existing_profile.save()
                    updated_count += 1
                else:
                    # Try to find by name within the same class
                    existing_user = User.objects.filter(
                        last_name=last_name, first_name=first_name, is_student=True
                    ).select_related('student_profile').first()

                    if existing_user and hasattr(existing_user, 'student_profile'):
                        existing_user.birth_date = birth_date
                        existing_user.save()
                        existing_user.student_profile.school_class = school_class
                        if personal_file_number:
                            existing_user.student_profile.personal_file_number = personal_file_number
                        existing_user.student_profile.save()
                        updated_count += 1
                    else:
                        password = generate_password()
                        user = User(
                            first_name=first_name,
                            last_name=last_name,
                            birth_date=birth_date,
                            is_student=True,
                            must_change_password=True,
                            temp_password=password,
                        )
                        user.set_password(password)
                        user.save()
                        StudentProfile.objects.create(
                            user=user,
                            school_class=school_class,
                            personal_file_number=personal_file_number,
                        )
                        created_count += 1

            except Exception as exc:
                errors.append(f'Лист «{sheet_name}», строка {row_idx} ({fio}): {exc}')

    return {
        'created': created_count,
        'updated': updated_count,
        'errors': errors,
    }
