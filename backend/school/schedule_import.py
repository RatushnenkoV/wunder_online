"""
Schedule import from Excel files.

Supported formats:
  - Classes file: rows = (day, period), cols = classes with (subject, [group2_subject], room)
  - Teachers file: rows = (day, period), cols = teachers with (subject, room)

Matching: teacher lessons are linked to class lessons by (weekday, period, room).
"""
import openpyxl
from io import BytesIO

from .models import GradeLevel, SchoolClass, Room, Subject, ScheduleLesson, ClassGroup, ClassSubject

DAY_MAP = {
    'понедельник': 1,
    'вторник': 2,
    'среда': 3,
    'четверг': 4,
    'пятница': 5,
}


def _normalize(s):
    return str(s).lower().strip() if s else ''


def _parse_class_name(raw):
    """Parse '1а' or '11б' → (number, letter_upper)."""
    raw = str(raw).strip()
    num = ''
    for i, c in enumerate(raw):
        if c.isdigit():
            num += c
        else:
            letter = raw[i:].strip().upper()
            return int(num), letter
    return int(num), ''


def parse_classes_file(file_bytes):
    """Parse the schedule-by-classes Excel file.

    Returns list of lesson dicts:
        class_name, weekday (1-5), period (int),
        subject_name, subject2_name (or None), room_name (or None),
        teacher_name (None — filled later by match_teachers)
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    max_col = ws.max_column

    header = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    # Detect class column layout (0-indexed into header/row lists)
    # Pattern A — 2-col: [class_name, 'каб']
    # Pattern B — 3-col: [class_name, None, 'каб']  (group split possible)
    class_cols = []
    i = 2  # skip col0 (day) and col1 (period)
    while i < len(header):
        val = header[i]
        if val is None:
            i += 1
            continue
        if _normalize(val) == 'каб':
            i += 1
            continue

        next1 = header[i + 1] if i + 1 < len(header) else None
        next2 = header[i + 2] if i + 2 < len(header) else None

        if next1 is not None and _normalize(next1) == 'каб':
            # 2-col class
            class_cols.append({
                'name': str(val).strip(),
                'subj_col': i,
                'subj2_col': None,
                'room_col': i + 1,
            })
            i += 2
        elif next1 is None and next2 is not None and _normalize(next2) == 'каб':
            # 3-col class (possible group split)
            class_cols.append({
                'name': str(val).strip(),
                'subj_col': i,
                'subj2_col': i + 1,
                'room_col': i + 2,
            })
            i += 3
        else:
            i += 1

    lessons = []
    current_day = None

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]

        day_val = _normalize(row[0])
        if day_val in DAY_MAP:
            current_day = DAY_MAP[day_val]

        period_val = row[1]
        if period_val is None or current_day is None:
            continue
        try:
            period = int(str(period_val).strip())
        except (ValueError, TypeError):
            continue

        for cls in class_cols:
            def _get(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            subj = _get(cls['subj_col'])
            subj2 = _get(cls['subj2_col'])
            room = _get(cls['room_col'])

            if not subj or str(subj).strip() == '':
                continue

            subj_str = str(subj).strip()
            subj2_str = str(subj2).strip() if subj2 else None
            room_str = str(room).strip() if room else None

            # If group subjects are the same — treat as single lesson
            if subj2_str == subj_str:
                subj2_str = None

            # For group lessons split comma-separated rooms: "303, 103" → room1=303, room2=103
            room1_str = room_str
            room2_str = None
            if subj2_str and room_str and ',' in room_str:
                parts = [r.strip() for r in room_str.split(',', 1)]
                room1_str = parts[0]
                room2_str = parts[1] if len(parts) > 1 else None

            lessons.append({
                'class_name': cls['name'],
                'weekday': current_day,
                'period': period,
                'subject_name': subj_str,
                'subject2_name': subj2_str,
                'room_name': room1_str,
                'room2_name': room2_str,   # only set for group lessons
                'teacher_name': None,
                'teacher2_name': None,     # filled by match_teachers
            })

    return lessons


def parse_teachers_file(file_bytes):
    """Parse the schedule-by-teachers Excel file.

    Returns list of teacher lesson dicts:
        teacher_name, weekday, period, subject_name, room_name
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    max_col = ws.max_column

    header = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    teacher_cols = []
    i = 2
    while i < len(header):
        val = header[i]
        if val is None:
            i += 1
            continue
        next1 = header[i + 1] if i + 1 < len(header) else None
        if next1 is not None and _normalize(next1) == 'каб':
            teacher_cols.append({
                'name': str(val).strip(),
                'subj_col': i,
                'room_col': i + 1,
            })
            i += 2
        else:
            i += 1

    teacher_lessons = []
    current_day = None

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]

        day_val = _normalize(row[0])
        if day_val in DAY_MAP:
            current_day = DAY_MAP[day_val]

        period_val = row[1]
        if period_val is None or current_day is None:
            continue
        try:
            period = int(str(period_val).strip())
        except (ValueError, TypeError):
            continue

        for t in teacher_cols:
            subj = row[t['subj_col']] if t['subj_col'] < len(row) else None
            room = row[t['room_col']] if t['room_col'] < len(row) else None

            if not subj or str(subj).strip() == '':
                continue

            teacher_lessons.append({
                'teacher_name': t['name'],
                'weekday': current_day,
                'period': period,
                'subject_name': str(subj).strip(),
                'room_name': str(room).strip() if room else None,
            })

    return teacher_lessons


def match_teachers(class_lessons, teacher_lessons):
    """Fill teacher_name (and teacher2_name for group lessons) in class_lessons
    by matching on (weekday, period, room). Modifies class_lessons in-place.
    """
    # Build index: (weekday, period, normalized_room_part) → teacher_name
    index = {}
    for tl in teacher_lessons:
        if not tl['room_name']:
            continue
        parts = [r.strip() for r in tl['room_name'].split(',')]
        for part in parts:
            key = (tl['weekday'], tl['period'], _normalize(part))
            if key not in index:
                index[key] = tl['teacher_name']

    for lesson in class_lessons:
        # Match teacher for group 1 by room_name
        if lesson.get('room_name'):
            parts = [r.strip() for r in lesson['room_name'].split(',')]
            for part in parts:
                key = (lesson['weekday'], lesson['period'], _normalize(part))
                if key in index:
                    lesson['teacher_name'] = index[key]
                    break
        # Match teacher for group 2 by room2_name
        if lesson.get('room2_name'):
            key = (lesson['weekday'], lesson['period'], _normalize(lesson['room2_name']))
            if key in index:
                lesson['teacher2_name'] = index[key]


def analyze(class_lessons, all_excel_teacher_names):
    """Compare Excel data against DB and return missing entities + DB options.

    all_excel_teacher_names: set of teacher name strings from the teachers file header
    """
    from accounts.models import User

    # Collect unique names from Excel
    excel_classes = sorted({l['class_name'] for l in class_lessons})
    excel_rooms_raw = []
    for l in class_lessons:
        if l.get('room_name'):
            excel_rooms_raw.append(l['room_name'])
        if l.get('room2_name'):
            excel_rooms_raw.append(l['room2_name'])

    # Expand comma-separated rooms
    excel_rooms = set()
    for raw in excel_rooms_raw:
        for part in raw.split(','):
            part = part.strip()
            if part:
                excel_rooms.add(part)
    excel_rooms = sorted(excel_rooms)

    # Teacher names: from header + from matched lessons (both groups)
    excel_teachers = set(all_excel_teacher_names)
    for l in class_lessons:
        if l.get('teacher_name'):
            excel_teachers.add(l['teacher_name'])
        if l.get('teacher2_name'):
            excel_teachers.add(l['teacher2_name'])
    excel_teachers = sorted(excel_teachers)

    # Load DB data
    db_classes_qs = list(
        SchoolClass.objects.select_related('grade_level')
        .values('id', 'grade_level__number', 'letter')
    )
    db_teachers_qs = list(
        User.objects.filter(is_teacher=True).values('id', 'first_name', 'last_name')
    )
    db_rooms_qs = list(Room.objects.values('id', 'name'))

    # Normalized lookup sets
    db_class_normalized = {}  # '1а' → id
    for c in db_classes_qs:
        key = f"{c['grade_level__number']}{c['letter'].lower()}"
        db_class_normalized[key] = c['id']

    db_teacher_by_last = {}  # last_name_lower → [user_dict, ...]
    for t in db_teachers_qs:
        key = t['last_name'].lower()
        db_teacher_by_last.setdefault(key, []).append(t)

    db_room_normalized = {}  # name_lower → id
    for r in db_rooms_qs:
        db_room_normalized[r['name'].lower().strip()] = r['id']

    # Find missing classes
    missing_classes = []
    for cls_name in excel_classes:
        normalized = cls_name.lower().replace(' ', '')
        if normalized not in db_class_normalized:
            missing_classes.append(cls_name)

    # Find missing teachers (with similar suggestions)
    missing_teachers = []
    for teacher_name in excel_teachers:
        parts = teacher_name.split()
        excel_last = parts[0].lower() if parts else ''
        # Check exact last name match
        if excel_last not in db_teacher_by_last:
            # Find similar by first 4 chars of last name
            similar = []
            for db_last, users in db_teacher_by_last.items():
                min_len = min(len(excel_last), len(db_last), 4)
                if min_len >= 2 and excel_last[:min_len] == db_last[:min_len]:
                    for u in users:
                        similar.append({
                            'id': u['id'],
                            'name': f"{u['last_name']} {u['first_name']}",
                        })
            missing_teachers.append({
                'name': teacher_name,
                'similar': similar,
            })

    # Find missing rooms
    missing_rooms = []
    for room_name in excel_rooms:
        if room_name.lower().strip() not in db_room_normalized:
            missing_rooms.append(room_name)

    return {
        'missing_classes': missing_classes,
        'missing_teachers': missing_teachers,
        'missing_rooms': missing_rooms,
        'db_classes': [
            {'id': c['id'], 'name': f"{c['grade_level__number']}-{c['letter']}"}
            for c in db_classes_qs
        ],
        'db_teachers': [
            {'id': t['id'], 'name': f"{t['last_name']} {t['first_name']}"}
            for t in db_teachers_qs
        ],
        'db_rooms': [
            {'id': r['id'], 'name': r['name']}
            for r in db_rooms_qs
        ],
    }


def execute_import(class_lessons, class_mappings, teacher_mappings, room_mappings, replace_existing):
    """Execute the actual import.

    class_mappings:   {excel_name: db_id or None}   None → create
    teacher_mappings: {excel_name: {'action': 'create'|'link'|'skip', 'id': db_id}}
    room_mappings:    {excel_name: db_id or None}   None → create
    replace_existing: bool — delete all ScheduleLessons before import
    """
    import random
    import string
    from accounts.models import User
    from .models import TeacherProfile

    if replace_existing:
        ScheduleLesson.objects.all().delete()

    # Resolve classes
    class_id_map = {}  # excel_name (lowercase) → db_id or None (skip)
    for excel_name, db_id in class_mappings.items():
        if db_id is None:
            # Create
            try:
                number, letter = _parse_class_name(excel_name)
            except Exception:
                continue
            grade, _ = GradeLevel.objects.get_or_create(number=number)
            sc, _ = SchoolClass.objects.get_or_create(grade_level=grade, letter=letter)
            class_id_map[excel_name.lower()] = sc.id
        else:
            class_id_map[excel_name.lower()] = int(db_id)

    # Resolve rooms
    # room_mappings value can be:
    #   None or {'action': 'create', 'name': '...'}  → create with given name
    #   int or {'action': 'link', 'id': N}            → link to existing
    room_id_map = {}  # normalized_excel_name → db_id
    for excel_name, mapping in room_mappings.items():
        norm_key = excel_name.lower().strip()
        if isinstance(mapping, dict):
            if mapping.get('action') == 'create':
                room_name = (mapping.get('name') or excel_name).strip()
                room, _ = Room.objects.get_or_create(name=room_name)
                room_id_map[norm_key] = room.id
            elif mapping.get('action') == 'link':
                room_id_map[norm_key] = int(mapping['id'])
        elif mapping is None:
            room, _ = Room.objects.get_or_create(name=excel_name.strip())
            room_id_map[norm_key] = room.id
        else:
            room_id_map[norm_key] = int(mapping)

    # Resolve teachers
    # teacher_mappings value:
    #   {'action': 'create', 'first_name': '...', 'last_name': '...'}
    #   {'action': 'link', 'id': N}
    #   {'action': 'skip'}
    teacher_id_map = {}  # excel_name → db_id or None (no teacher)
    for excel_name, mapping in teacher_mappings.items():
        action = mapping.get('action', 'skip')
        if action == 'link':
            teacher_id_map[excel_name] = mapping.get('id')
        elif action == 'create':
            # Use frontend-supplied names; fallback to parsing excel_name
            first_name = (mapping.get('first_name') or '').strip()
            last_name = (mapping.get('last_name') or '').strip()
            if not first_name and not last_name:
                parts = excel_name.strip().split()
                last_name = parts[0] if parts else excel_name
                first_name = parts[1] if len(parts) > 1 else 'Н/А'
            elif not first_name:
                first_name = 'Н/А'
            elif not last_name:
                last_name = first_name
                first_name = 'Н/А'
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
            user = User.objects.create_user(
                first_name, last_name, password,
                is_teacher=True,
                must_change_password=True,
            )
            TeacherProfile.objects.get_or_create(user=user)
            teacher_id_map[excel_name] = user.id
        else:
            teacher_id_map[excel_name] = None

    # Cache for ClassGroup pairs per class (created on demand for group lessons)
    class_group_cache = {}  # class_id → (group1, group2)

    def get_class_groups(class_id):
        if class_id not in class_group_cache:
            g1, _ = ClassGroup.objects.get_or_create(school_class_id=class_id, name='Группа 1')
            g2, _ = ClassGroup.objects.get_or_create(school_class_id=class_id, name='Группа 2')
            class_group_cache[class_id] = (g1, g2)
        return class_group_cache[class_id]

    # Cache to avoid duplicate ClassSubject lookups
    class_subject_seen = set()  # (class_id, name_lower)

    def ensure_class_subject(class_id, name):
        key = (class_id, name.lower())
        if key in class_subject_seen:
            return
        class_subject_seen.add(key)
        if not ClassSubject.objects.filter(school_class_id=class_id, name=name).exists():
            try:
                ClassSubject.objects.create(school_class_id=class_id, name=name)
            except Exception:
                pass  # already created by concurrent request or unique violation

    # Pre-load DB caches for fallback lookups (entities already in DB, not in mappings)
    db_room_cache = {r.name.lower().strip(): r.id for r in Room.objects.all()}
    db_teacher_cache = {}  # last_name_lower → user_id (first match)
    for t in User.objects.filter(is_teacher=True):
        key = t.last_name.lower()
        if key not in db_teacher_cache:
            db_teacher_cache[key] = t.id

    _MISSING = object()

    def resolve_room(room_raw):
        """Return DB room id for a raw room string; tries mapping then DB."""
        if not room_raw:
            return None
        for part in [r.strip() for r in room_raw.split(',')]:
            norm = part.lower().strip()
            rid = room_id_map.get(norm)
            if rid:
                return rid
            rid = db_room_cache.get(norm)
            if rid:
                return rid
        return None

    def resolve_teacher(teacher_name):
        """Return DB user id for a teacher name string; tries mapping then DB."""
        if not teacher_name:
            return None
        # Try explicit mapping first (handles missing teachers resolved in UI)
        # Use sentinel to distinguish 'skip' (value=None) from 'not in mapping'
        tid = teacher_id_map.get(teacher_name, _MISSING)
        if tid is not _MISSING:
            return tid  # None means explicitly skipped
        # Fallback: look up in DB by last name (teacher already existed in DB)
        parts = teacher_name.strip().split()
        last = parts[0].lower() if parts else ''
        return db_teacher_cache.get(last)

    # Import lessons
    created = 0
    skipped = 0
    errors = []

    for lesson in class_lessons:
        excel_class = lesson['class_name']
        class_id = class_id_map.get(excel_class.lower())
        if class_id is None:
            # Not in mappings → try to find in DB directly
            try:
                number, letter = _parse_class_name(excel_class)
                sc = SchoolClass.objects.filter(
                    grade_level__number=number, letter__iexact=letter
                ).first()
                if sc:
                    class_id = sc.id
            except Exception:
                pass
            if class_id is None:
                skipped += 1
                continue

        subj_name = lesson.get('subject_name', '').strip()
        if not subj_name:
            skipped += 1
            continue

        subject, _ = Subject.objects.get_or_create(name=subj_name)
        room_id = resolve_room(lesson.get('room_name'))
        teacher_id = resolve_teacher(lesson.get('teacher_name'))
        subject2_name = lesson.get('subject2_name')

        # For group lessons: get/create ClassGroup pair and assign group1 to first lesson
        group1_id = None
        group2 = None
        if subject2_name:
            group1, group2 = get_class_groups(class_id)
            group1_id = group1.id

        try:
            ScheduleLesson.objects.create(
                school_class_id=class_id,
                weekday=lesson['weekday'],
                lesson_number=lesson['period'],
                subject=subject,
                teacher_id=teacher_id,
                room_id=room_id,
                group_id=group1_id,
            )
            created += 1
        except Exception as e:
            errors.append(f"{excel_class} {lesson['weekday']}/{lesson['period']} {subj_name}: {e}")
            continue

        ensure_class_subject(class_id, subj_name)

        # Handle group split (different subject for group 2)
        if subject2_name and group2:
            subject2, _ = Subject.objects.get_or_create(name=subject2_name)
            room2_id = resolve_room(lesson.get('room2_name'))
            teacher2_id = resolve_teacher(lesson.get('teacher2_name'))
            try:
                ScheduleLesson.objects.create(
                    school_class_id=class_id,
                    weekday=lesson['weekday'],
                    lesson_number=lesson['period'],
                    subject=subject2,
                    teacher_id=teacher2_id,
                    room_id=room2_id,
                    group_id=group2.id,
                )
                created += 1
            except Exception as e:
                errors.append(f"{excel_class} group2 {lesson['weekday']}/{lesson['period']}: {e}")
                continue
            ensure_class_subject(class_id, subject2_name)

    return {
        'created': created,
        'skipped': skipped,
        'errors': errors[:20],  # limit to first 20 errors
    }
